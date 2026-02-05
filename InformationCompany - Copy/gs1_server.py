import os
import datetime
import json
import urllib.parse
from typing import Dict, Any

# Excel support (optional)
try:
    from openpyxl import Workbook, load_workbook

    EXCEL_AVAILABLE = True
except Exception:
    EXCEL_AVAILABLE = False

# HTML parsing support (optional)
try:
    from bs4 import BeautifulSoup

    BS4_AVAILABLE = True
except Exception:
    BS4_AVAILABLE = False

from flask import Flask, request, jsonify, make_response, send_file
import requests
import time
import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

app = Flask(__name__)


# Simple CORS without extra dependency
@app.after_request
def add_cors_headers(response):
    origin = request.headers.get("Origin", "*")
    response.headers["Access-Control-Allow-Origin"] = origin if origin else "*"
    response.headers["Vary"] = "Origin"
    response.headers["Access-Control-Allow-Methods"] = "GET, POST, OPTIONS"
    response.headers["Access-Control-Allow-Headers"] = "Content-Type"
    response.headers["Access-Control-Allow-Credentials"] = "true"
    return response


@app.route("/", methods=["GET"])
def root():
    ui_path = os.path.join(os.path.dirname(__file__), "gs1_ui.html")
    if os.path.exists(ui_path):
        return send_file(ui_path)
    return "<h1>GS1 Backend API Running</h1><p>Place gs1_ui.html in the same folder</p>", 200


# Avoid noisy 404s from browser favicon requests
@app.route("/favicon.ico")
def favicon():
    return ("", 204)


@app.route("/api/search", methods=["POST", "OPTIONS"])
def api_search():
    if request.method == "OPTIONS":
        return make_response("", 204)

    try:
        data: Dict[str, Any] = request.get_json(force=True) or {}
    except Exception:
        return jsonify({"ok": False, "error": "Invalid JSON body"}), 400

    company_name = (data.get("company_name") or "").strip()
    country = (data.get("country") or "VN").strip()
    street_address = (data.get("street_address") or "").strip()
    city = (data.get("city") or "").strip()
    addresses = data.get("addresses") or []

    # Validate - at least one search parameter
    if not (company_name or street_address or city or (isinstance(addresses, list) and len(addresses) > 0)):
        return jsonify({
            "ok": False,
            "error": "Vui lòng cung cấp ít nhất một trong: company_name, street_address, city"
        }), 400

    base_url = "https://www.gs1.org/services/verified-by-gs1/results"
    full_url = base_url
    queries = []
    seen = set()
    if street_address or city:
        q = {
            "company_name": company_name,
            "country": country,
            "street_address": street_address,
            "city": city,
        }
        key = (q["street_address"].lower(), q["city"].lower())
        if key not in seen:
            queries.append(q)
            seen.add(key)
    if isinstance(addresses, list):
        for a in addresses:
            sa = (a.get("street_address") or "").strip()
            ci = (a.get("city") or "").strip()
            if not (sa or ci):
                continue
            q = {
                "company_name": company_name,
                "country": country,
                "street_address": sa,
                "city": ci,
            }
            key = (q["street_address"].lower(), q["city"].lower())
            if key not in seen:
                queries.append(q)
                seen.add(key)
    if not queries:
        queries.append({
            "company_name": company_name,
            "country": country,
            "street_address": "",
            "city": "",
        })

    # Use undetected-chromedriver to bypass CAPTCHA detection
    options = uc.ChromeOptions()

    headless_env = os.getenv("HEADLESS", "1").strip()

    SHOW_BROWSER = False
    if not SHOW_BROWSER and headless_env != "0":
        options.add_argument("--headless=new")

    # Tùy chọn để giảm khả năng bị phát hiện
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--no-sandbox")
    options.add_argument("--window-size=1920,1200")
    options.add_argument("--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36")

    # Disable images for faster loading
    prefs = {
        "profile.managed_default_content_settings.images": 2,
        "profile.default_content_setting_values.notifications": 2,
    }
    options.add_experimental_option("prefs", prefs)

    extracted_rows = []
    status_code = 200

    try:
        print("[INFO] Đang khởi động undetected Chrome...")
        driver = uc.Chrome(options=options, version_main=None)
        print("[INFO] ✅ Đã khởi động trình duyệt")

        debug_mode = os.getenv("DEBUG", "0") == "1"

        def try_click_any(selectors, by="xpath", wait_sec=5):
            for sel in selectors:
                try:
                    if by == "xpath":
                        elem = WebDriverWait(driver, wait_sec).until(
                            EC.element_to_be_clickable((By.XPATH, sel))
                        )
                    else:
                        elem = WebDriverWait(driver, wait_sec).until(
                            EC.element_to_be_clickable((By.CSS_SELECTOR, sel))
                        )
                    elem.click()
                    time.sleep(0.3)
                    print(f"[INFO] ✅ Đã click: {sel[:60]}...")
                    return True
                except Exception:
                    continue
            return False

        for q in queries:
            query_string = urllib.parse.urlencode({k: v for k, v in q.items() if v})
            full_url = f"{base_url}?{query_string}" if query_string else base_url
            print(f"[INFO] Đang truy cập: {full_url}")
            driver.get(full_url)
            time.sleep(2)

            if debug_mode:
                try:
                    screenshot_path = os.path.join(os.path.dirname(__file__), "debug_screenshot.png")
                    driver.save_screenshot(screenshot_path)
                    print(f"[DEBUG] Screenshot saved: {screenshot_path}")
                except:
                    pass

            print("[INFO] Đang đợi Cookie popup xuất hiện...")
            try:
                WebDriverWait(driver, 5).until(
                    EC.presence_of_element_located((By.XPATH, "//button[contains(text(), 'Accept')]"))
                )
                time.sleep(0.5)
            except:
                pass

            cookie_clicked = try_click_any([
                "//button[contains(text(), 'Accept all optional cookies')]",
                "//button[text()='Accept all optional cookies']",
                "//button[contains(@class, 'ot-sdk-button') and contains(text(), 'Accept')]",
                "//button[contains(text(), 'Reject all optional cookies')]",
                "//button[@id='onetrust-accept-btn-handler']",
                "//button[contains(text(), 'Accept all')]",
                "//button[contains(@class, 'cookie') and contains(text(), 'Accept')]",
            ], by="xpath", wait_sec=3)

            if not cookie_clicked:
                try:
                    driver.execute_script("""
                        var cookieBanners = document.querySelectorAll('[class*="cookie"], [id*="cookie"], [class*="onetrust"], [id*="onetrust"]');
                        cookieBanners.forEach(function(el) { el.style.display = 'none'; el.remove(); });
                    """)
                except Exception:
                    pass

            print("[INFO] Đang đợi Terms popup...")
            try:
                WebDriverWait(driver, 4).until(
                    EC.presence_of_element_located((By.XPATH, "//div[@role='dialog']//button | //button[contains(text(), 'I agree')]"))
                )
                time.sleep(0.5)
            except:
                pass

            terms_clicked = try_click_any([
                "//button[text()='Accept']",
                "//button[contains(text(), 'I agree')]",
                "//div[@role='dialog']//button[1]",
            ], by="xpath", wait_sec=3)

            if not terms_clicked:
                try:
                    driver.execute_script("""
                        var dialogs = document.querySelectorAll('[role="dialog"], .ui-dialog, .modal, .popup');
                        dialogs.forEach(function(el) { el.style.display = 'none'; el.remove(); });
                    """)
                except Exception:
                    pass

            captcha_solved = False
            has_captcha = False
            try:
                WebDriverWait(driver, 2).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "iframe[src*='recaptcha']"))
                )
                has_captcha = True
                time.sleep(1)
            except:
                pass

            if captcha_solved or not has_captcha:
                print("[INFO] Click nút Search...")
                _ = try_click_any([
                    "//button[contains(@class, 'button--primary') and contains(., 'Search')]",
                    "//form//button[@type='submit']",
                    "//button[contains(@class,'btn') and contains(., 'Search')]",
                ], by="xpath", wait_sec=3)

            print("[INFO] Đang tăng số kết quả trên trang...")
            try:
                dropdown = WebDriverWait(driver, 4).until(
                    EC.element_to_be_clickable((By.XPATH, "//select[contains(@class, 'page-size') or contains(@name, 'items_per_page')]"))
                )
                dropdown.click()
                time.sleep(0.2)
                option_20 = driver.find_element(By.XPATH, "//option[@value='20' or text()='20']")
                option_20.click()
                time.sleep(1)
            except Exception:
                pass

            time.sleep(1)
            selectors_to_try = [
                ".views-table tbody tr",
                "table tbody tr",
                ".view-content table tbody tr",
                ".view-content .views-table tr",
                "tbody tr",
                "tr.views-row",
                ".view-gs1-verified tbody tr",
                "table.views-view-grid tbody tr",
                ".views-view-grid tbody tr",
                ".table tbody tr",
                "table tr"
            ]

            rows = []
            for selector in selectors_to_try:
                try:
                    WebDriverWait(driver, 12).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, selector))
                    )
                    rows = driver.find_elements(By.CSS_SELECTOR, selector)
                    if rows:
                        print(f"[INFO] ✅ Tìm thấy {len(rows)} hàng với: {selector}")
                        break
                except Exception:
                    continue

            if not rows:
                try:
                    page_html = driver.page_source
                    if any(msg in page_html for msg in ["No results", "No results found", "Không tìm thấy", "0 results"]):
                        print("[INFO] No results for this query")
                    else:
                        print("[INFO] Không tìm thấy bảng dữ liệu phù hợp")
                except Exception:
                    pass
                continue

            for r in rows:
                tds = r.find_elements(By.TAG_NAME, "td")
                if len(tds) >= 4:
                    extracted_rows.append({
                        "Licence Key": (tds[0].text or tds[0].get_attribute("textContent") or "").strip(),
                        "Company Name": (tds[1].text or tds[1].get_attribute("textContent") or "").strip(),
                        "City": (tds[2].text or tds[2].get_attribute("textContent") or "").strip(),
                        "Country": (tds[3].text or tds[3].get_attribute("textContent") or "").strip(),
                    })

            try:
                next_pages = 0
                max_pages = 10
                while next_pages < max_pages:
                    try:
                        next_btn = driver.find_element(By.XPATH, "//a[@rel='next' and not(contains(@class, 'disabled'))] | //a[contains(@class, 'pager__item--next') and not(contains(@class, 'disabled'))]")
                        if next_btn:
                            next_btn.click()
                            time.sleep(2)
                            new_rows = driver.find_elements(By.CSS_SELECTOR, selectors_to_try[0] if selectors_to_try else "table tbody tr")
                            for r in new_rows:
                                tds = r.find_elements(By.TAG_NAME, "td")
                                if len(tds) >= 4:
                                    row_data = {
                                        "Licence Key": (tds[0].text or tds[0].get_attribute("textContent") or "").strip(),
                                        "Company Name": (tds[1].text or tds[1].get_attribute("textContent") or "").strip(),
                                        "City": (tds[2].text or tds[2].get_attribute("textContent") or "").strip(),
                                        "Country": (tds[3].text or tds[3].get_attribute("textContent") or "").strip(),
                                    }
                                    if row_data not in extracted_rows:
                                        extracted_rows.append(row_data)
                            next_pages += 1
                        else:
                            break
                    except:
                        break
            except Exception:
                pass

        print(f"[INFO] 🎉 Hoàn thành! Tổng cộng đã trích xuất {len(extracted_rows)} hàng dữ liệu")

    except Exception as e:
        error_msg = str(e)
        print(f"[ERROR] Lỗi scrape: {error_msg}")
        try:
            driver.quit()
        except Exception:
            pass
        return jsonify({
            "ok": False,
            "url": full_url,
            "error": f"Lỗi khi scrape: {error_msg[:200]}",
            "hint": "Thử HEADLESS=0 để xem trình duyệt, hoặc tăng thời gian chờ.",
        }), 500
    finally:
        try:
            driver.quit()
        except Exception:
            pass

    if not extracted_rows:
        try:
            last_html = page_html if 'page_html' in locals() else ''
        except Exception:
            last_html = ''
        return jsonify({
            "ok": False,
            "url": full_url,
            "status_code": status_code,
            "error": "Không tìm thấy kết quả hoặc cấu trúc trang thay đổi",
            "hint": "Thử HEADLESS=0 để quan sát, hoặc thay đổi từ khóa tìm kiếm.",
            "html_preview": last_html[:800] if last_html else None
        }), 200

    # Ensure data directory
    data_dir = os.path.join(os.path.dirname(__file__), "data")
    os.makedirs(data_dir, exist_ok=True)

    # Write directly to Excel (Extracted sheet only)
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_file = os.path.join(data_dir, "results.xlsx")
    excel_written = False
    excel_note = None
    extracted_written = False

    if EXCEL_AVAILABLE:
        try:
            # Retry logic nếu file đang bị lock
            max_retries = 3
            retry_count = 0
            wb = None

            while retry_count < max_retries:
                try:
                    # Load hoặc tạo workbook
                    if os.path.exists(excel_file):
                        wb = load_workbook(excel_file)
                        print(f"[INFO] Đã load Excel hiện có: {excel_file}")
                    else:
                        wb = Workbook()
                        # Xóa sheet mặc định nếu có
                        if "Sheet" in wb.sheetnames:
                            wb.remove(wb["Sheet"])
                        print(f"[INFO] Tạo Excel file mới: {excel_file}")
                    break  # Thành công, thoát loop

                except PermissionError:
                    retry_count += 1
                    if retry_count < max_retries:
                        print(f"[WARNING] File Excel đang được mở. Retry {retry_count}/{max_retries}...")
                        time.sleep(2)
                    else:
                        raise Exception(
                            f"Không thể mở file Excel sau {max_retries} lần thử. "
                            f"Vui lòng ĐÓNG Excel nếu đang mở file '{excel_file}' và thử lại!"
                        )

            if not wb:
                raise Exception("Không thể tạo workbook")

            sheet_name = "Extracted"

            # Tạo sheet nếu chưa có
            if sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(sheet_name)
                ws.append(["timestamp", "Licence Key", "Company Name", "City", "Country"])
                print(f"[INFO] Tạo sheet mới: {sheet_name}")
            else:
                ws = wb[sheet_name]
                print(f"[INFO] Append vào sheet hiện có: {sheet_name}")

            # Đếm số dòng hiện tại (không kể header)
            current_rows = ws.max_row - 1

            # Lấy danh sách Licence Key đã có để check trùng
            existing_company_names = set()
            for row in ws.iter_rows(min_row=2, min_col=3, max_col=3, values_only=True):
                if row[0]:  # Cột C (Company Name)
                    existing_company_names.add(str(row[0]).strip().lower())

            print(f"[INFO] Đã có {len(existing_company_names)} Company Name trong Excel")

            # Thêm dữ liệu (append, không ghi đè) và check trùng
            new_rows_added = 0
            duplicates_skipped = 0

            for row in extracted_rows:
                company_name = str(row.get("Company Name", "")).strip().lower()

                # Kiểm tra trùng
                if company_name and company_name in existing_company_names:
                    duplicates_skipped += 1
                    print(f"[INFO] ⏭️  Bỏ qua trùng Company Name: {row.get('Company Name', '')[:50]}")
                    continue

                # Thêm vào Excel
                ws.append([
                    ts,
                    str(row.get("Licence Key", "")).strip(),
                    row.get("Company Name", ""),
                    row.get("City", ""),
                    row.get("Country", "")
                ])

                # Đánh dấu đã thêm
                if company_name:
                    existing_company_names.add(company_name)
                new_rows_added += 1

            # Retry khi save
            retry_count = 0
            while retry_count < max_retries:
                try:
                    wb.save(excel_file)
                    excel_written = True
                    extracted_written = bool(extracted_rows)
                    print(f"[INFO] ✅ Đã lưu Excel:")
                    print(f"       - Dòng trước đó: {current_rows}")
                    print(f"       - Dòng thêm mới: {new_rows_added}")
                    print(f"       - Tổng dòng hiện tại: {ws.max_row - 1}")
                    print(f"       - File: {excel_file}")
                    break

                except PermissionError:
                    retry_count += 1
                    if retry_count < max_retries:
                        print(f"[WARNING] Không thể ghi file (đang mở?). Retry {retry_count}/{max_retries}...")
                        time.sleep(2)
                    else:
                        raise Exception(
                            f"Không thể lưu file Excel sau {max_retries} lần thử. "
                            f"Vui lòng ĐÓNG Excel và thử lại!"
                        )

        except Exception as e:
            excel_note = f"Lỗi khi ghi Excel: {e}"
            print(f"[ERROR] {excel_note}")
    else:
        excel_note = "openpyxl chưa được cài đặt. Chạy: pip install openpyxl"

    return jsonify({
        "ok": True,
        "url": full_url,
        "status_code": status_code,
        "excel_file": excel_file,
        "excel_written": excel_written,
        "excel_note": excel_note,
        "extracted_count": len(extracted_rows),
        "extracted_written": extracted_written,
        "extracted_rows": extracted_rows,  # Trả về TẤT CẢ dữ liệu
        "note": f"✅ Đã quét và lưu {len(extracted_rows)} hàng vào Excel (sheet Extracted).",
    })


if __name__ == "__main__":
    # PORT can be customized via env var if needed
    port = int(os.getenv("PORT", "5000"))
    print(f"\n{'=' * 60}")
    print(f"🚀 GS1 Backend API đang chạy tại: http://127.0.0.1:{port}")
    print(f"📁 File Excel sẽ được lưu tại: ./data/results.xlsx")
    print(f"🔍 Debug mode: HEADLESS={os.getenv('HEADLESS', '1')} (0=hiển thị browser, 1=ẩn)")
    print(f"{'=' * 60}\n")
    app.run(host="127.0.0.1", port=port, debug=True)