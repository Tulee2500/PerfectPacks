import os
import time
import argparse
import urllib.parse

import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from zipfile import BadZipFile


def build_url(company_name: str, country: str, street_address: str, city: str) -> str:
    base = "https://www.gs1.org/services/verified-by-gs1/results"
    params = {
        "company_name": company_name or "",
        "country": country or "VN",
        "street_address": street_address or "",
        "city": city or "",
    }
    qs = urllib.parse.urlencode({k: v for k, v in params.items() if v})
    return f"{base}?{qs}" if qs else base


def ensure_data_dir(project_root: str) -> str:
    data_dir = os.path.join(project_root, "data")
    os.makedirs(data_dir, exist_ok=True)
    return data_dir


def append_df_to_excel(file_path: str, df: pd.DataFrame, sheet_name: str = "Extracted") -> None:
    # Normalize and keep only required columns in order
    df = df.rename(columns={"License Key": "Licence Key"})
    cols = ["Licence Key", "Company Name", "City", "Country"]
    df = df[cols]

    # If target file doesn't exist or is empty, create fresh workbook
    if not os.path.exists(file_path) or os.path.getsize(file_path) == 0:
        df.to_excel(file_path, index=False, sheet_name=sheet_name)
        return

    try:
        with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
            # Ensure workbook can be loaded (handles some pandas versions that lazy-load)
            writer.book = load_workbook(file_path)
            if sheet_name not in writer.book.sheetnames:
                df.to_excel(writer, index=False, sheet_name=sheet_name)
            else:
                ws = writer.book[sheet_name]
                start_row = ws.max_row + 1
                df.to_excel(
                    writer,
                    index=False,
                    sheet_name=sheet_name,
                    header=False,
                    startrow=start_row - 1,
                )
    except (BadZipFile, InvalidFileException):
        # Existing file is not a valid XLSX (e.g., corrupted or wrong format). Backup and recreate.
        backup_path = file_path + ".bak"
        try:
            if os.path.exists(backup_path):
                os.remove(backup_path)
            os.replace(file_path, backup_path)
        except Exception:
            pass
        df.to_excel(file_path, index=False, sheet_name=sheet_name)


def scrape(company_name: str, country: str, street_address: str, city: str, headless: bool) -> pd.DataFrame:
    options = Options()
    if headless:
        options.add_argument("--headless=new")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--start-maximized")
    options.add_argument("--window-size=1920,1200")

    driver = webdriver.Chrome(options=options)
    url = build_url(company_name, country, street_address, city)
    driver.get(url)

    try:
        # Accept cookie banner if present
        try:
            cookie_accept = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.ID, "onetrust-accept-btn-handler"))
            )
            cookie_accept.click()
            time.sleep(0.5)
        except Exception:
            pass

        # Accept Terms of Use if present (selector may change)
        try:
            terms_btn = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//button[contains(., 'Accept') or contains(., 'I agree')][1]"))
            )
            terms_btn.click()
            time.sleep(0.5)
        except Exception:
            pass

        # Wait for any table rows to appear
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "table tbody tr, .views-table tbody tr"))
        )

        rows = driver.find_elements(By.CSS_SELECTOR, "table tbody tr, .views-table tbody tr")
        data = []
        for row in rows:
            cols = row.find_elements(By.TAG_NAME, "td")
            if len(cols) >= 4:
                data.append({
                    "License Key": cols[0].text.strip(),
                    "Company Name": cols[1].text.strip(),
                    "City": cols[2].text.strip(),
                    "Country": cols[3].text.strip(),
                })

        return pd.DataFrame(data)
    finally:
        time.sleep(1)
        driver.quit()


def main():
    parser = argparse.ArgumentParser(description="Scrape GS1 results and append to Excel")
    parser.add_argument("--company", dest="company_name", required=True, help="Company name keyword")
    parser.add_argument("--address", dest="street_address", default="", help="Street address keyword")
    parser.add_argument("--city", dest="city", default="", help="City keyword")
    parser.add_argument("--country", dest="country", default="VN", help="Country code, default VN")
    parser.add_argument("--headless", action="store_true", help="Run Chrome headless")

    args = parser.parse_args()

    project_root = os.path.dirname(__file__)
    data_dir = ensure_data_dir(project_root)
    excel_path = os.path.join(data_dir, "results.xlsx")

    df = scrape(args.company_name, args.country, args.street_address, args.city, args.headless)
    if df.empty:
        print("No rows scraped. Check if the page structure changed or add more wait.")
        return

    append_df_to_excel(excel_path, df, sheet_name="Extracted")
    print(f"Appended {len(df)} rows to: {excel_path}")


if __name__ == "__main__":
    main()