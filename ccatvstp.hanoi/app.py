from flask import Flask, request, jsonify, send_from_directory, send_file
from flask_cors import CORS
import requests
from bs4 import BeautifulSoup
import re
from urllib.parse import urlparse, parse_qs, urlencode, urlunparse
import time
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__, static_folder='.')
CORS(app)


def get_url_with_page(base_url, page_num):
    """Thêm số trang vào URL"""
    parsed = urlparse(base_url)
    params = parse_qs(parsed.query)
    params['page'] = [str(page_num)]
    new_query = urlencode(params, doseq=True)
    return urlunparse((parsed.scheme, parsed.netloc, parsed.path, parsed.params, new_query, parsed.fragment))


def remove_duplicates(companies):
    """Loại bỏ các công ty trùng lặp, chỉ giữ bản ghi đầu tiên"""
    seen = set()
    unique_companies = []

    for company in companies:
        name = company.get('ten_doanh_nghiep', '')
        if name and name not in seen:
            seen.add(name)
            unique_companies.append(company)

    return unique_companies


def check_duplicates(companies):
    """Kiểm tra các công ty trùng lặp"""
    from collections import Counter

    company_names = [c['ten_doanh_nghiep'] for c in companies if c.get('ten_doanh_nghiep')]
    name_counts = Counter(company_names)

    # Lấy các công ty xuất hiện nhiều hơn 1 lần
    duplicates = []
    for name, count in name_counts.items():
        if count > 1:
            # Tìm tất cả bản ghi của công ty này
            records = [c for c in companies if c.get('ten_doanh_nghiep') == name]
            duplicates.append({
                'ten_cong_ty': name,
                'so_lan_xuat_hien': count,
                'san_pham': [r.get('ten_san_pham', '') for r in records]
            })

    # Sắp xếp theo số lần xuất hiện giảm dần
    duplicates.sort(key=lambda x: x['so_lan_xuat_hien'], reverse=True)

    return duplicates


def scrape_single_page(url):
    """Scrape dữ liệu từ một trang"""
    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
            'Accept-Language': 'vi-VN,vi;q=0.9,en-US;q=0.8,en;q=0.7',
            'Accept-Encoding': 'gzip, deflate',
            'Connection': 'keep-alive',
        }

        response = requests.get(url, headers=headers, timeout=30)
        response.encoding = 'utf-8'

        if response.status_code != 200:
            return None, f'Status code: {response.status_code}'

        soup = BeautifulSoup(response.content, 'html.parser')

        # Tìm bảng
        table = None
        table = soup.find('table', class_='table table-bordered bg-white resize-column')
        if not table:
            table = soup.find('table', class_='table-bordered')
        if not table:
            tables = soup.find_all('table')
            for t in tables:
                if t.find('tbody'):
                    table = t
                    break

        if not table:
            main_body = soup.find('div', class_='main-body')
            if main_body:
                table_content = main_body.find('div', class_='table-content')
                if table_content:
                    table = table_content.find('table')

        if not table:
            return None, 'Không tìm thấy bảng dữ liệu'

        tbody = table.find('tbody')
        if not tbody:
            return None, 'Không tìm thấy tbody'

        rows = tbody.find_all('tr')
        companies = []

        for idx, row in enumerate(rows):
            try:
                company_data = {}
                cells = row.find_all('td')

                if len(cells) == 0:
                    continue

                # STT
                stt_cell = row.find('td', class_='STT')
                if stt_cell:
                    company_data['stt'] = stt_cell.text.strip()
                elif len(cells) > 0:
                    company_data['stt'] = cells[0].text.strip()

                # Tên doanh nghiệp
                enterprise_cell = row.find('td', class_='enterprisesInfo_name')
                if enterprise_cell:
                    name_span = enterprise_cell.find('span', class_='enterprise-name')
                    company_data[
                        'ten_doanh_nghiep'] = name_span.text.strip() if name_span else enterprise_cell.text.strip()
                elif len(cells) > 1:
                    company_data['ten_doanh_nghiep'] = cells[1].text.strip()

                # Địa chỉ
                address_cell = row.find('td', class_='enterprisesInfo_full_address')
                if address_cell:
                    company_data['dia_chi'] = address_cell.text.strip()
                elif len(cells) > 2:
                    company_data['dia_chi'] = cells[2].text.strip()

                # Tên sản phẩm
                product_cell = row.find('td', class_='productInfo_name')
                if product_cell:
                    company_data['ten_san_pham'] = product_cell.text.strip()
                elif len(cells) > 3:
                    company_data['ten_san_pham'] = cells[3].text.strip()

                # Mã hồ sơ
                record_cell = row.find('td', class_='record_no')
                company_data['ma_ho_so'] = record_cell.text.strip() if record_cell else ''

                # Nhóm sản phẩm
                group_cell = row.find('td', class_='productGroup_name')
                if group_cell:
                    company_data['nhom_san_pham'] = group_cell.text.strip()
                elif len(cells) > 4:
                    company_data['nhom_san_pham'] = cells[4].text.strip()

                # Ngày công bố
                date_cell = row.find('td', class_='productInfo_publish_date')
                if date_cell:
                    company_data['ngay_cong_bo'] = date_cell.text.strip()
                elif len(cells) > 5:
                    company_data['ngay_cong_bo'] = cells[5].text.strip()

                # Ghi chú
                comment_cell = row.find('td', class_='record_comment_result')
                company_data['ghi_chu'] = comment_cell.text.strip() if comment_cell else ''

                companies.append(company_data)

            except Exception as e:
                print(f"⚠️ Lỗi khi xử lý dòng {idx + 1}: {str(e)}")
                continue

        # Lấy tổng số trang
        pagination = soup.find('div', class_='pagination-tracuu')
        total_pages = 1

        if pagination:
            page_items = pagination.find_all('li', class_='page-item')
            for item in page_items:
                link = item.find('a', class_='page-link')
                if link:
                    page_num_text = link.text.strip()
                    if page_num_text.isdigit():
                        total_pages = max(total_pages, int(page_num_text))

        return companies, total_pages

    except Exception as e:
        return None, str(e)


def scrape_company_data(url, start_page=1, end_page=1):
    """Scrape dữ liệu công ty từ URL với nhiều trang"""
    try:
        print(f"🔍 Đang trích xuất từ trang {start_page} đến trang {end_page}")

        all_companies = []
        total_pages = end_page

        for page_num in range(start_page, end_page + 1):
            # Tạo URL cho trang hiện tại
            if page_num == 1 and 'page=' not in url:
                current_url = url
            else:
                current_url = get_url_with_page(url, page_num)

            print(f"📄 Đang scrape trang {page_num}/{end_page}...")

            companies, detected_total_pages = scrape_single_page(current_url)

            if companies is None:
                print(f"❌ Lỗi tại trang {page_num}: {detected_total_pages}")
                continue

            if isinstance(detected_total_pages, int):
                total_pages = max(total_pages, detected_total_pages)

            all_companies.extend(companies)
            print(f"✅ Trang {page_num}: Tìm thấy {len(companies)} bản ghi")

            # Delay giữa các request để tránh bị block
            if page_num < end_page:
                time.sleep(1)

        print(f"✅ Hoàn thành! Tổng cộng: {len(all_companies)} bản ghi từ {end_page - start_page + 1} trang")

        # Kiểm tra trùng lặp
        duplicates = check_duplicates(all_companies)
        unique_count = len(set(c['ten_doanh_nghiep'] for c in all_companies if c.get('ten_doanh_nghiep')))

        # Loại bỏ trùng lặp - chỉ giữ bản ghi đầu tiên
        unique_companies = remove_duplicates(all_companies)

        return {
            'success': True,
            'data': all_companies,
            'data_unique': unique_companies,
            'total_records': len(all_companies),
            'unique_records': len(unique_companies),
            'removed_duplicates': len(all_companies) - len(unique_companies),
            'unique_companies': unique_count,
            'duplicates': duplicates,
            'start_page': start_page,
            'end_page': end_page,
            'total_pages': total_pages,
            'pages_scraped': end_page - start_page + 1
        }

    except Exception as e:
        print(f"❌ Lỗi: {str(e)}")
        import traceback
        traceback.print_exc()
        return {'error': f'Lỗi: {str(e)}'}


@app.route('/')
def index():
    """Serve index.html"""
    return send_from_directory('.', 'index.html')


@app.route('/api/scrape', methods=['POST'])
def scrape():
    """API endpoint để scrape dữ liệu"""
    data = request.get_json()
    url = data.get('url', '')
    start_page = data.get('start_page', 1)
    end_page = data.get('end_page', 1)

    if not url:
        return jsonify({'error': 'URL không được để trống'}), 400

    if not url.startswith('http'):
        return jsonify({'error': 'URL không hợp lệ'}), 400

    try:
        start_page = int(start_page)
        end_page = int(end_page)
    except:
        return jsonify({'error': 'Số trang không hợp lệ'}), 400

    if start_page < 1 or end_page < 1:
        return jsonify({'error': 'Số trang phải lớn hơn 0'}), 400

    if start_page > end_page:
        return jsonify({'error': 'Trang bắt đầu phải nhỏ hơn hoặc bằng trang kết thúc'}), 400

    if end_page - start_page > 50:
        return jsonify({'error': 'Chỉ được scrape tối đa 50 trang mỗi lần'}), 400

    result = scrape_company_data(url, start_page, end_page)

    # Tự động tạo file Excel
    if result.get('success'):
        try:
            excel_file = create_excel_report(result, url, start_page, end_page)
            result['excel_file'] = excel_file
            result['excel_created'] = True
        except Exception as e:
            print(f"⚠️ Không thể tạo file Excel: {str(e)}")
            result['excel_created'] = False

    return jsonify(result)


@app.route('/api/test', methods=['GET'])
def test():
    """Test endpoint"""
    return jsonify({
        'status': 'OK',
        'message': 'Server đang chạy bình thường'
    })


@app.route('/download/<filename>')
def download_file(filename):
    """Download file Excel"""
    try:
        file_path = os.path.join('exports', filename)
        if os.path.exists(file_path):
            return send_file(file_path, as_attachment=True)
        else:
            return jsonify({'error': 'File không tồn tại'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500


def create_excel_report(data, url, start_page, end_page):
    """Tạo file Excel với nhiều sheet"""

    # Tạo thư mục exports nếu chưa có
    if not os.path.exists('exports'):
        os.makedirs('exports')

    # Tạo tên file
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'bao_cao_cong_ty_{timestamp}.xlsx'
    filepath = os.path.join('exports', filename)

    # Tạo workbook
    wb = Workbook()

    # Style định dạng
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # === SHEET 1: Thống kê ===
    ws_stats = wb.active
    ws_stats.title = "Thống kê"

    ws_stats['A1'] = 'BÁO CÁO TRÍCH XUẤT DỮ LIỆU CÔNG TY'
    ws_stats['A1'].font = Font(bold=True, size=16, color='4472C4')
    ws_stats.merge_cells('A1:B1')

    ws_stats['A3'] = 'Thời gian tạo:'
    ws_stats['B3'] = datetime.now().strftime('%d/%m/%Y %H:%M:%S')

    ws_stats['A4'] = 'URL nguồn:'
    ws_stats['B4'] = url[:100] + '...' if len(url) > 100 else url

    ws_stats['A5'] = 'Phạm vi trang:'
    ws_stats['B5'] = f'Từ trang {start_page} đến trang {end_page}'

    ws_stats['A7'] = 'THỐNG KÊ'
    ws_stats['A7'].font = Font(bold=True, size=14)

    stats_data = [
        ['Tổng số bản ghi:', data.get('total_records', 0)],
        ['Bản ghi duy nhất:', data.get('unique_records', 0)],
        ['Bản ghi trùng lặp:', data.get('removed_duplicates', 0)],
        ['Số công ty duy nhất:', data.get('unique_companies', 0)],
        ['Số trang đã scrape:', data.get('pages_scraped', 0)],
    ]

    for i, (label, value) in enumerate(stats_data, start=8):
        ws_stats[f'A{i}'] = label
        ws_stats[f'B{i}'] = value
        ws_stats[f'A{i}'].font = Font(bold=True)

    ws_stats.column_dimensions['A'].width = 30
    ws_stats.column_dimensions['B'].width = 50

    # === SHEET 2: Dữ liệu Unique ===
    ws_unique = wb.create_sheet("Dữ liệu Unique")
    headers = ['STT', 'Tên doanh nghiệp', 'Địa chỉ', 'Tên sản phẩm', 'Mã hồ sơ', 'Nhóm sản phẩm', 'Ngày công bố',
               'Ghi chú']

    for col, header in enumerate(headers, start=1):
        cell = ws_unique.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    unique_data = data.get('data_unique', [])
    for row_idx, item in enumerate(unique_data, start=2):
        ws_unique.cell(row=row_idx, column=1, value=item.get('stt', ''))
        ws_unique.cell(row=row_idx, column=2, value=item.get('ten_doanh_nghiep', ''))
        ws_unique.cell(row=row_idx, column=3, value=item.get('dia_chi', ''))
        ws_unique.cell(row=row_idx, column=4, value=item.get('ten_san_pham', ''))
        ws_unique.cell(row=row_idx, column=5, value=item.get('ma_ho_so', ''))
        ws_unique.cell(row=row_idx, column=6, value=item.get('nhom_san_pham', ''))
        ws_unique.cell(row=row_idx, column=7, value=item.get('ngay_cong_bo', ''))
        ws_unique.cell(row=row_idx, column=8, value=item.get('ghi_chu', ''))

        for col in range(1, 9):
            ws_unique.cell(row=row_idx, column=col).border = border

    # Tự động điều chỉnh độ rộng cột
    ws_unique.column_dimensions['A'].width = 8
    ws_unique.column_dimensions['B'].width = 35
    ws_unique.column_dimensions['C'].width = 50
    ws_unique.column_dimensions['D'].width = 40
    ws_unique.column_dimensions['E'].width = 25
    ws_unique.column_dimensions['F'].width = 20
    ws_unique.column_dimensions['G'].width = 15
    ws_unique.column_dimensions['H'].width = 20

    # === SHEET 3: Tất cả dữ liệu ===
    ws_all = wb.create_sheet("Tất cả dữ liệu")
    for col, header in enumerate(headers, start=1):
        cell = ws_all.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    all_data = data.get('data', [])
    for row_idx, item in enumerate(all_data, start=2):
        ws_all.cell(row=row_idx, column=1, value=item.get('stt', ''))
        ws_all.cell(row=row_idx, column=2, value=item.get('ten_doanh_nghiep', ''))
        ws_all.cell(row=row_idx, column=3, value=item.get('dia_chi', ''))
        ws_all.cell(row=row_idx, column=4, value=item.get('ten_san_pham', ''))
        ws_all.cell(row=row_idx, column=5, value=item.get('ma_ho_so', ''))
        ws_all.cell(row=row_idx, column=6, value=item.get('nhom_san_pham', ''))
        ws_all.cell(row=row_idx, column=7, value=item.get('ngay_cong_bo', ''))
        ws_all.cell(row=row_idx, column=8, value=item.get('ghi_chu', ''))

        for col in range(1, 9):
            ws_all.cell(row=row_idx, column=col).border = border

    ws_all.column_dimensions['A'].width = 8
    ws_all.column_dimensions['B'].width = 35
    ws_all.column_dimensions['C'].width = 50
    ws_all.column_dimensions['D'].width = 40
    ws_all.column_dimensions['E'].width = 25
    ws_all.column_dimensions['F'].width = 20
    ws_all.column_dimensions['G'].width = 15
    ws_all.column_dimensions['H'].width = 20

    # === SHEET 4: Danh sách trùng lặp ===
    duplicates = data.get('duplicates', [])
    if duplicates:
        ws_dup = wb.create_sheet("Công ty trùng lặp")

        dup_headers = ['STT', 'Tên công ty', 'Số lần xuất hiện', 'Danh sách sản phẩm']
        for col, header in enumerate(dup_headers, start=1):
            cell = ws_dup.cell(row=1, column=col, value=header)
            cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        for row_idx, dup in enumerate(duplicates, start=2):
            ws_dup.cell(row=row_idx, column=1, value=row_idx - 1)
            ws_dup.cell(row=row_idx, column=2, value=dup.get('ten_cong_ty', ''))
            ws_dup.cell(row=row_idx, column=3, value=dup.get('so_lan_xuat_hien', 0))
            ws_dup.cell(row=row_idx, column=4, value=', '.join(dup.get('san_pham', [])))

            for col in range(1, 5):
                ws_dup.cell(row=row_idx, column=col).border = border

        ws_dup.column_dimensions['A'].width = 8
        ws_dup.column_dimensions['B'].width = 40
        ws_dup.column_dimensions['C'].width = 20
        ws_dup.column_dimensions['D'].width = 60

    # Lưu file
    wb.save(filepath)
    print(f"✅ Đã tạo file Excel: {filename}")

    return filename


if __name__ == '__main__':
    print("=" * 60)
    print("🚀 Server đang chạy tại: http://127.0.0.1:777")
    print("📝 Mở trình duyệt và truy cập địa chỉ trên để sử dụng")
    print("=" * 60)
    app.run(debug=True, host='0.0.0.0', port=777)