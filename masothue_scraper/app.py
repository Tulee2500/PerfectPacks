from flask import Flask, request, jsonify, send_file
from scraper import search_companies_batch
import pandas as pd
import io
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB


@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        if 'file' not in request.files:
            return jsonify({"error": "Không có file được upload"}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({"error": "Vui lòng chọn file"}), 400

        if not (file.filename.endswith('.csv') or file.filename.endswith('.xlsx')):
            return jsonify({"error": "Chỉ hỗ trợ file .csv hoặc .xlsx"}), 400

        try:
            # Đọc file
            if file.filename.endswith('.csv'):
                encodings = ['utf-8-sig', 'utf-8', 'cp1258', 'latin1']
                file.stream.seek(0)
                raw = file.stream.read()
                df = None
                for enc in encodings:
                    try:
                        text = raw.decode(enc, errors='replace')
                        df = pd.read_csv(io.StringIO(text), dtype=str, keep_default_na=False)
                        break
                    except:
                        continue
                if df is None:
                    return jsonify({"error": "Không đọc được file CSV"}), 400
            else:
                file.stream.seek(0)
                df = pd.read_excel(file, engine='openpyxl', dtype=str)

            # Lấy cột đầu tiên
            company_names = df.iloc[:, 0].astype(str).str.strip()
            company_names = company_names[company_names != ''].dropna().tolist()

            if not company_names:
                return jsonify({"error": "File không chứa dữ liệu"}), 400

            # Tìm kiếm
            print(f"\n🔍 Bắt đầu tìm kiếm {len(company_names)} công ty...")
            results = search_companies_batch(company_names)

            # Đếm thống kê
            found = sum(1 for r in results if r['status'] == 'found')
            not_found = sum(1 for r in results if r['status'] == 'not_found')
            error = sum(1 for r in results if r['status'] == 'error')

            return jsonify({
                "success": True,
                "total": len(company_names),
                "found": found,
                "not_found": not_found,
                "error": error,
                "results": results
            })

        except Exception as e:
            return jsonify({"error": f"Lỗi xử lý: {str(e)}"}), 400

    # HTML form
    return """
<!DOCTYPE html>
<html lang="vi">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Tra cứu Mã số thuế hàng loạt</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body {
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
            display: flex;
            align-items: center;
            justify-content: center;
            padding: 20px;
        }
        .container {
            background: white;
            border-radius: 20px;
            box-shadow: 0 20px 60px rgba(0,0,0,0.3);
            padding: 40px;
            max-width: 800px;
            width: 100%;
        }
        h1 {
            color: #667eea;
            text-align: center;
            margin-bottom: 30px;
            font-size: 2rem;
        }
        .upload-area {
            border: 3px dashed #667eea;
            border-radius: 15px;
            padding: 60px 20px;
            text-align: center;
            cursor: pointer;
            transition: all 0.3s;
            background: #f8f9ff;
        }
        .upload-area:hover {
            background: #e8eaff;
            border-color: #764ba2;
        }
        .upload-area.dragover {
            background: #d8dcff;
            border-color: #764ba2;
        }
        .upload-icon {
            font-size: 4rem;
            margin-bottom: 20px;
        }
        .btn {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            border: none;
            padding: 15px 40px;
            border-radius: 10px;
            font-size: 1.1rem;
            font-weight: 600;
            cursor: pointer;
            width: 100%;
            margin-top: 20px;
            transition: transform 0.2s;
        }
        .btn:hover { transform: translateY(-2px); }
        .btn:disabled {
            opacity: 0.6;
            cursor: not-allowed;
        }
        .file-name {
            margin-top: 20px;
            padding: 15px;
            background: #e8f5e9;
            border-radius: 10px;
            color: #2e7d32;
            font-weight: 600;
            display: none;
        }
        .results {
            margin-top: 30px;
            display: none;
        }
        .stats {
            display: grid;
            grid-template-columns: repeat(3, 1fr);
            gap: 15px;
            margin-bottom: 30px;
        }
        .stat-card {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 20px;
            border-radius: 10px;
            text-align: center;
        }
        .stat-number {
            font-size: 2.5rem;
            font-weight: 700;
            margin-bottom: 5px;
        }
        .stat-label {
            font-size: 0.9rem;
            opacity: 0.9;
        }
        table {
            width: 100%;
            border-collapse: collapse;
            margin-top: 20px;
            font-size: 0.9rem;
        }
        th {
            background: #667eea;
            color: white;
            padding: 12px 8px;
            text-align: left;
            font-weight: 600;
            position: sticky;
            top: 0;
        }
        td {
            padding: 10px 8px;
            border-bottom: 1px solid #eee;
        }
        tr:hover {
            background: #f5f5f5;
        }
        .status-found {
            background: #d4edda;
            color: #155724;
            padding: 4px 8px;
            border-radius: 5px;
            font-size: 0.85rem;
            font-weight: 600;
        }
        .status-not-found {
            background: #f8d7da;
            color: #721c24;
            padding: 4px 8px;
            border-radius: 5px;
            font-size: 0.85rem;
            font-weight: 600;
        }
        .table-container {
            max-height: 500px;
            overflow-y: auto;
            border-radius: 10px;
            border: 1px solid #ddd;
        }
        .info {
            background: #e3f2fd;
            border-left: 4px solid #2196f3;
            padding: 15px;
            border-radius: 5px;
            margin-top: 20px;
        }
        .info h3 {
            color: #1976d2;
            font-size: 1.1rem;
            margin-bottom: 10px;
        }
        .info ul {
            margin-left: 20px;
            line-height: 1.8;
        }
    </style>
</head>
<body>
    <div class="container">
        <h1>🔍 Tra cứu Mã số thuế hàng loạt</h1>

        <div class="upload-area" id="uploadArea">
            <div class="upload-icon">📁</div>
            <h2>Kéo thả file hoặc click để chọn</h2>
            <p style="color: #666; margin-top: 10px;">Hỗ trợ: Excel (.xlsx), CSV (.csv)</p>
            <input type="file" id="fileInput" accept=".xlsx,.csv" style="display: none;">
        </div>

        <div class="file-name" id="fileName"></div>

        <button class="btn" id="submitBtn" disabled>🚀 Bắt đầu tìm kiếm</button>

        <div class="results" id="results">
            <h2 style="color: #667eea; margin-bottom: 20px;">📊 Kết quả</h2>

            <div class="stats">
                <div class="stat-card">
                    <div class="stat-number" id="foundCount">0</div>
                    <div class="stat-label">Tìm thấy</div>
                </div>
                <div class="stat-card">
                    <div class="stat-number" id="notFoundCount">0</div>
                    <div class="stat-label">Không tìm thấy</div>
                </div>
                <div class="stat-card">
                    <div class="stat-number" id="errorCount">0</div>
                    <div class="stat-label">Lỗi</div>
                </div>
            </div>

            <div class="table-container">
                <table id="resultsTable">
                    <thead>
                        <tr>
                            <th style="width: 5%;">STT</th>
                            <th style="width: 15%;">Tìm kiếm</th>
                            <th style="width: 20%;">Tên công ty</th>
                            <th style="width: 12%;">MST</th>
                            <th style="width: 25%;">Địa chỉ</th>
                            <th style="width: 11%;">Điện thoại</th>
                            <th style="width: 12%;">Trạng thái</th>
                        </tr>
                    </thead>
                    <tbody id="resultsBody"></tbody>
                </table>
            </div>

            <button class="btn" id="exportBtn" style="margin-top: 20px;">📥 Tải xuống Excel</button>
        </div>

        <div class="info">
            <h3>💡 Hướng dẫn sử dụng</h3>
            <ul>
                <li>Tạo file Excel hoặc CSV với danh sách tên công ty ở cột đầu tiên</li>
                <li>Upload file và click "Bắt đầu tìm kiếm"</li>
                <li>Hệ thống sẽ tự động tra cứu từ masothue.com</li>
                <li>Kết quả có thể tải xuống định dạng Excel</li>
            </ul>
        </div>
    </div>

    <script>
        const uploadArea = document.getElementById('uploadArea');
        const fileInput = document.getElementById('fileInput');
        const fileName = document.getElementById('fileName');
        const submitBtn = document.getElementById('submitBtn');
        const results = document.getElementById('results');
        const exportBtn = document.getElementById('exportBtn');

        let currentResults = [];

        uploadArea.addEventListener('click', () => fileInput.click());

        uploadArea.addEventListener('dragover', (e) => {
            e.preventDefault();
            uploadArea.classList.add('dragover');
        });

        uploadArea.addEventListener('dragleave', () => {
            uploadArea.classList.remove('dragover');
        });

        uploadArea.addEventListener('drop', (e) => {
            e.preventDefault();
            uploadArea.classList.remove('dragover');
            fileInput.files = e.dataTransfer.files;
            updateFileName();
        });

        fileInput.addEventListener('change', updateFileName);

        function updateFileName() {
            if (fileInput.files.length > 0) {
                fileName.textContent = `✅ ${fileInput.files[0].name}`;
                fileName.style.display = 'block';
                submitBtn.disabled = false;
            }
        }

        submitBtn.addEventListener('click', async () => {
            if (!fileInput.files.length) return;

            const formData = new FormData();
            formData.append('file', fileInput.files[0]);

            submitBtn.disabled = true;
            submitBtn.textContent = '⏳ Đang tìm kiếm...';
            results.style.display = 'none';

            try {
                const response = await fetch('/', {
                    method: 'POST',
                    body: formData
                });

                const data = await response.json();

                if (data.success) {
                    currentResults = data.results;
                    displayResults(data);
                } else {
                    alert('Lỗi: ' + data.error);
                }
            } catch (error) {
                alert('Lỗi: ' + error.message);
            } finally {
                submitBtn.disabled = false;
                submitBtn.textContent = '🚀 Bắt đầu tìm kiếm';
            }
        });

        function displayResults(data) {
            document.getElementById('foundCount').textContent = data.found;
            document.getElementById('notFoundCount').textContent = data.not_found;
            document.getElementById('errorCount').textContent = data.error;

            const tbody = document.getElementById('resultsBody');
            tbody.innerHTML = '';

            data.results.forEach((result, index) => {
                const row = document.createElement('tr');

                let statusHtml = '';
                if (result.status === 'found') {
                    statusHtml = '<span class="status-found">✅ Tìm thấy</span>';
                } else if (result.status === 'not_found') {
                    statusHtml = '<span class="status-not-found">❌ Không tìm</span>';
                } else {
                    statusHtml = '<span class="status-not-found">⚠️ Lỗi</span>';
                }

                row.innerHTML = `
                    <td>${index + 1}</td>
                    <td><strong>${result.company_name}</strong></td>
                    <td>${result.name || '-'}</td>
                    <td><code>${result.tax_code || '-'}</code></td>
                    <td title="${result.address || '-'}">${(result.address || '-').substring(0, 50)}${result.address && result.address.length > 50 ? '...' : ''}</td>
                    <td>${result.phone || '-'}</td>
                    <td>${statusHtml}</td>
                `;

                tbody.appendChild(row);
            });

            results.style.display = 'block';
        }

        exportBtn.addEventListener('click', async () => {
            try {
                const response = await fetch('/export', {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify({ results: currentResults })
                });

                const blob = await response.blob();
                const url = window.URL.createObjectURL(blob);
                const a = document.createElement('a');
                a.href = url;
                a.download = `masothue_${new Date().toISOString().slice(0,10)}.xlsx`;
                a.click();
                window.URL.revokeObjectURL(url);
            } catch (error) {
                alert('Lỗi export: ' + error.message);
            }
        });
    </script>
</body>
</html>
    """


@app.route("/export", methods=["POST"])
def export_excel():
    results = request.get_json().get("results", [])

    if not results:
        return jsonify({"error": "Không có dữ liệu"}), 400

    rows = []
    for idx, r in enumerate(results, 1):
        rows.append({
            "STT": idx,
            "Tìm kiếm": r.get("company_name", ""),
            "Tên công ty": r.get("name", "Không có"),
            "Mã số thuế": r.get("tax_code", "Không có"),
            "Địa chỉ": r.get("address", "Không có"),
            "Điện thoại": r.get("phone", "Không có"),
            "Người đại diện": r.get("representative", "Không có"),
            "Trạng thái": r.get("status", "unknown"),
            "URL": r.get("url", "")
        })

    df = pd.DataFrame(rows)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Kết quả")
        ws = writer.sheets["Kết quả"]

        # Style header
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Auto-fit columns
        for i, col in enumerate(df.columns, 1):
            max_len = max(df[col].astype(str).map(len).max(), len(col)) + 2
            ws.column_dimensions[get_column_letter(i)].width = min(max_len, 60)

        # Wrap text for address
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    output.seek(0)
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="masothue_results.xlsx"
    )


if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=888)