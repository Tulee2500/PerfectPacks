from flask import Flask, render_template, request, jsonify, send_file
from scraper_hsctvn import search_by_tax_code, search_companies_batch
import csv
import io
import sys
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Cấu hình console UTF-8
try:
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')
except Exception:
    pass

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB


@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        tax_code = request.form.get("tax_code", "").strip()
        if tax_code:
            data = search_by_tax_code(tax_code)

            if "error" in data:
                return render_template("index_hsctvn.html", error=data["error"])

            if "detail" in data:
                return render_template("detail_hsctvn.html", info=data["detail"])
            else:
                return render_template("index_hsctvn.html", error="Không tìm thấy công ty nào")

    return render_template("index_hsctvn.html")


@app.route("/batch", methods=["GET", "POST"])
def batch_search():
    """Tìm kiếm hàng loạt từ file Excel/CSV"""
    if request.method == "POST":
        if 'file' not in request.files:
            return render_template("batch_hsctvn.html", error="Không có file được upload"), 400

        file = request.files['file']
        if file.filename == '':
            return render_template("batch_hsctvn.html", error="Vui lòng chọn file"), 400

        if not (file.filename.endswith('.csv') or file.filename.endswith('.xlsx')):
            return render_template("batch_hsctvn.html", error="Chỉ hỗ trợ file .csv hoặc .xlsx"), 400

        try:
            # Đọc file
            if file.filename.endswith('.csv'):
                encodings = ['utf-8-sig', 'utf-8', 'cp1258', 'cp1252', 'latin1']
                file.stream.seek(0)
                raw = file.stream.read()
                df = None
                for enc in encodings:
                    try:
                        text = raw.decode(enc, errors='replace')
                        import io as _io
                        df = pd.read_csv(_io.StringIO(text), dtype=str, keep_default_na=False)
                        break
                    except:
                        continue
                if df is None:
                    raise Exception('Không đọc được file CSV')
            else:
                file.stream.seek(0)
                df = pd.read_excel(file, engine='openpyxl', dtype=str)

            # Lấy cột đầu tiên làm danh sách mã số thuế
            tax_codes = df.iloc[:, 0].astype(str).map(lambda x: x.strip()).replace(
                {'': None, 'nan': None}).dropna().tolist()

            if not tax_codes:
                return render_template("batch_hsctvn.html", error="File không chứa dữ liệu"), 400

            print(f"Bat dau tim kiem {len(tax_codes)} ma so thue...")
            results = search_companies_batch(tax_codes)

            return render_template("batch_results_hsctvn.html", results=results, total=len(tax_codes))

        except Exception as e:
            return render_template("batch_hsctvn.html", error=f"Lỗi xử lý file: {str(e)}"), 400

    return render_template("batch_hsctvn.html")


@app.route("/export_excel", methods=["POST"])
def export_results_excel():
    results = request.get_json().get("results", [])

    if not results:
        return jsonify({"error": "Không có dữ liệu"}), 400

    fieldnames = [
        "Mã số thuế tìm kiếm",
        "Tên công ty",
        "Điện thoại",
        "Địa chỉ"
    ]

    rows = []
    for result in results:
        rows.append({
            "Mã số thuế tìm kiếm": result.get("tax_code_search", ""),
            "Tên công ty": result.get("name", "Không có"),
            "Điện thoại": result.get("phone", "Không có"),
            "Địa chỉ": result.get("address", "Không có"),
        })

    df = pd.DataFrame(rows, columns=fieldnames)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Kết quả")
        ws = writer.sheets["Kết quả"]

        # Header styling
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="FFDDEBF7", end_color="FFDDEBF7", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center")

        # Auto-width columns
        for i, col in enumerate(df.columns, 1):
            maxlen = max([len(str(x)) if x is not None else 0 for x in [col] + df[col].tolist()])
            ws.column_dimensions[get_column_letter(i)].width = min(maxlen + 2, 60)

        # Wrap text for address and company name
        wrap_alignment = Alignment(wrap_text=True, vertical="top")
        addr_col = df.columns.get_loc("Địa chỉ") + 1
        name_col = df.columns.get_loc("Tên công ty") + 1

        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=addr_col).alignment = wrap_alignment
            ws.cell(row=r, column=name_col).alignment = wrap_alignment

        # Text format for tax search code and phone
        tax_search_col = df.columns.get_loc("Mã số thuế tìm kiếm") + 1
        phone_col = df.columns.get_loc("Điện thoại") + 1

        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=tax_search_col).number_format = "@"
            ws.cell(row=r, column=phone_col).number_format = "@"

    output.seek(0)
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="hsctvn_results.xlsx"
    )


if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5002)  # Port 5001 để tránh conflict